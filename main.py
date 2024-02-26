import pyautogui
import shutil
import asyncio
import sys
import os
import openpyxl # to handle input data through Excel
from openapp import (openCRM,connectToCRM)
from searchandupdateot import (searchAndNotacion)
from closeapps import (closeCRM)
from genericfunctions import (showDesktop,make_noise,get_username_os,sendEmail,getCurrentDateAndTime,terminateProcess)
from telegramfunctions import (sendTelegramMsg,sendTelegramMsgWithDocuments)
import logging
from logsfunctions import (setupLogger)
from datetime import datetime
from time import sleep
from dotenv import load_dotenv

async def main():   
    
    # while True:
    # print(pyautogui.position())    
    
    load_dotenv()  

    # CONSTANTS
    TELEGRAM_CHAT_ID = 5970685607 # PERSONAL CHAT WITH BOT 5970685607 # TELEGRAM CHAT GROUP ID #-1002019721248
    INPUT_DIRECTORY = 'C:/ActualizacionComentarios/Insumos/'
    INPUT_FILENAME = 'Input BOT 001 V1.xlsx'
    SUPER_LOG_FILENAME = 'C:/ActualizacionComentarios/logs/super_log.txt'
    GENERIC_ERROR_MSG = 'No Se ha procesado el incidente - Para ver mas detalles ver el archivo de logs'
    CORREOS_DESTINO = 'correos'
    CORREOS_CC = 'correo'

    # VARIABLES        
    counter = 1
    total_rows = 0
    total_cols = 0
    resp_funcion = None
    
    try:
        #*** SHOW MENU TO SELECT RPA ACTIVITY ******#
        
        print("----------------- BOT proceso de notaciones --------------------")                        
        print("Introduzca el numero de la actividad que desea ejecutar")
        actividad_rpa = input("1. Ingresar Notaciones\n2. No realizar actividad\n")
            
        if actividad_rpa == "1":
            actividad_rpa_selected = 'INGRESAR NOTACIONES OTP'

        else:
            print("Actividad no permitida!")            
            print("Exit")
            exit_program()

        #************* DATA EXTRACTION *************#
        
        file = os.path.join(INPUT_DIRECTORY,INPUT_FILENAME)
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        if 'Input Backup' in wb.sheetnames:
            del wb['Input Backup']          
        
        # Excel Input Base Manipulation
        target_ws = wb.copy_worksheet(ws)
        target_ws.title = "Input Backup"
        total_rows = len(ws['A'])
        total_cols = len(ws[1])
        print("Total de registros en base fuente: ",total_rows-1)
        print("Total de columnas en base fuente: ",total_cols)

        #************* OPEN APPS *************#
        #CRM        
        #openCRM()
        #sleep(1)
        #connectToCRM()   

        #************* DATA PROCESSING *************#
        
        rows = ws.iter_rows(min_row=2, max_row=total_rows, min_col=1, max_col=total_cols)    
        
        for row in rows:
            # se omiten aquellos registros que en la col STATUS tenga COMPLETADO o contenga la palabra ERROR
            if row[1].value == 'COMPLETADO' or str(row[1].value).find("ERROR") != -1:
                continue

            #Interacting with the CRM Depending on the user input call the corresponding method            
            if actividad_rpa.lower() == "1":
                resp_funcion = searchAndNotacion(str(row[0].value),row[3].value,row[4].value,row[5].value,row[6].value,row[7].value)  
                print("Respuesta de la función:",resp_funcion)         
            
            # Manejo de las posibles respuestas identificadas en los metodos de control
            if resp_funcion == 0:
                ot_completadas = [str(row[0].value)]
                row[1].value = 'COMPLETADO'
                # logging.warning('Se ha procesado el incidente: %s',row[0].value,extra=attrs)
            elif resp_funcion == 10:
                # logging.warning('No se identifica vista de detalles de OT reconocida en el CRM',extra=attrs)
                row[1].value = 'ERROR - No se identifica vista de detalles de OT reconocida en el CRM'                
            elif resp_funcion == 9:
                # logging.warning('Se presenta mensaje de advertencia en el proceso en CRM',extra=attrs)                
                row[1].value = 'ERROR - Se presenta mensaje de advertencia en el proceso en CRM'                    
            else:              
                row[1].value = 'ERROR - '+ GENERIC_ERROR_MSG
                # logging.warning('No Se ha procesado el incidente: %s',row[0].value,extra=attrs)                
                terminateProcess('CRM.exe')
                print("Se ha terminado el proceso CRM.exe")
                sleep(2)
                openCRM()            
                connectToCRM()
                
            print("Registro(counter) #: ",counter)
            if counter % 10 == 0:
                await sendTelegramMsg('Se han procesado '+ str(counter) +' registros y se procede a reiniciar CRM ',TELEGRAM_CHAT_ID)
                terminateProcess('CRM.exe')
                print("Se ha terminado el proceso CRM.exe")
                sleep(2)
                openCRM()
                connectToCRM()
                                            
            counter = counter + 1
            wb.save(file) 
                       
        print("Fin del Proceso Macro")  

    except FileNotFoundError as e:
        exit_program()
        
    except Exception as e:
        
        print(f"An error occurred: {e}")
        # logging.warning('Error Occurred: %s',e,extra=attrs)        
        e_type, e_object, e_traceback = sys.exc_info()
        e_line_number = e_traceback.tb_lineno
        print (sys.exc_info())
       
        exit_program()
    
    else:
        #************* CLOSE APPS  *************#
        closeCRM()
        sleep(1)
        showDesktop()                            

        #*********** SEND NOTIFICATIONS  *******#
        # Esto se ejecutara si el bloque try se ejecuta sin errores
        print("Try block succesfully executed")
        await sendTelegramMsg('END - RPA '+actividad_rpa_selected+' ha finalizado.\n' + 'Usuario ['+ get_username_os() +']\nTiempo de finalización:'+ getCurrentDateAndTime(),TELEGRAM_CHAT_ID)
        await sendTelegramMsgWithDocuments(TELEGRAM_CHAT_ID)             
        exit_program()        
    finally:
        sleep(1)
        make_noise() 
        
def exit_program():
    sys.exit(0)    
    
if __name__ == "__main__":
   asyncio.run(main())